'''
The application is intended to visualize and save data obtained from ADCP sensors and stored in Matlab (mat) format
in winADCP software. winADCP saves ADCP data in Matlab Level 1 format.
Module '_read_matlab1_AK' is used to read matlab file into single dictionary.
Application passes the dictionary to imported module '_fill_ADCP_mx' to parse the dictionary into several variables and
numpy arrays for easier management.

The application consists of 4 windows:
Main control implemented in class MainWindow and represents table view of ADCP data for loaded file. Multiple files may
be loaded and browsed through the window. There are options to export ADCP data in Excel format (either
loaded file or all available) and export all the loaded data (multiple files) into single GIS compatible GEOJSON or
WKT ASCII file as point with relevant attributes. Combobox is used to navigate through loaded files.

Profile view is implemented in class ProfileView and represents 3D OpenGL widget showing single ADCP profile along with
its data in numeric format. Mouse wheel is used to navigate through the window. Buttons 'BWD'/'FWD' are used to
navigate through profiles. View may be set in orthographic or perspective projection by 'Ortho'/'Persp' buttons.
Current image may be saved in png format by pressing 'Save this image' button. 'Save all images' button is used to
save all the profiles of the loaded file with the same settings. Text entry is used to type text to be plotted on the
top of legend of the exported images.

Compass view is implemented in class CompassView. It parents subclass CompassWidget. The view represents vertical
compass view of the current and neighbouring profiles along with relevant metadata (timestamp, Lat, Lon).
The view is centered on the current profile. Number of visible profiles may be changed by mouse wheel. View limits and
vector scale may be set using relevant text entries. Buttons 'BWD'/'FWD' are used to navigate through profiles.
Current image may be saved in png format by pressing 'Save this image' button. 'Save all images' button is used to save
all the profiles of the loaded file with the same settings. Text entry is used to type text to be plotted on the top
of legend of the exported images.

Map view is implemented in class GeoView. It represents map view of all the loaded ADCP data. Vector scale may be set
using scale text entry. Slider is used to navigate through bins (depths). 'Animate depth' checkbox used to turn on/off
background color with depth. Current image may be saved in png format by pressing 'Save this image' button. Text entry
is used to type text to be plotted on the top of legend of the exported image.
'''


import os
from pathlib import Path
import math
import sys
import json
import numpy as np
from PIL import Image, ImageDraw, ImageFont
from PySide6 import QtWidgets, QtGui, QtCore
from PySide6.QtGui import QIntValidator, QDoubleValidator
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import QFileDialog, QAbstractItemView, QMessageBox, QSizePolicy
import pyqtgraph as pg
import pyqtgraph.opengl as gl
from pyqtgraph import Vector
import pyqtgraph.exporters
from openpyxl import Workbook
import _UI_control_3D_Geo
import _UI_profview
import _UI_compassview
import _UI_geoview
import _UI_license
import _UI_help
import _read_matlab1_AK
import _fill_ADCP_mx


OPTIONS = QFileDialog.Options()


class LicenseWin(QtWidgets.QMainWindow, _UI_license.Ui_MainWindow):
    '''
    Shows license window
    '''
    def __init__(self):
        super().__init__()
        self.setupUi(self)


class HelpWin(QtWidgets.QMainWindow, _UI_help.Ui_MainWindow):
    '''
    Shows help window
    '''
    def __init__(self):
        super().__init__()
        self.setupUi(self)


class InvertArrow(pg.ArrowItem):
    '''
    This class is used to override pg ArrowItem class with arrow base offset (from tip to bottom)
    '''
    def paint(self, p, *args):
        p.translate(-2 * self.boundingRect().center())
        pg.ArrowItem.paint(self, p, *args)


class CompassWidget(pg.PlotWidget):
    '''
    This class is used to explicitly enable wheelEvent access and override 2D wheel scrolling with 1D (horizontal)
    '''
    def __init__(self):
        super().__init__()

        # self.getViewBox().invertY(True)
        self.setMinimumSize(525, 580)
        self.setMaximumSize(100000, 100000)

        self.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Maximum)

        self.hideButtons()                                      # hide A(autozoom) button
        self.plotItem.setMenuEnabled(False)                     # Disable context menu
        self.getPlotItem().setMouseEnabled(x=False, y=False)    # Disable mouse panning & zooming - !wheelEvent overrides this


    def wheelEvent(self, e):
        # print(e.angleDelta().y())
        cv.view_win = cv.view_win - 2 if e.angleDelta().y() < 0 else cv.view_win + 2
        cv.view_win = 0 if cv.view_win < 0 else cv.view_win
        cv.drawcompass(mc.col_to_show)

    # def mouseReleaseEvent(self, e):
    #     print("mouse released")
    #
    # def mousePressEvent(self, e):
    #     print("mousepressed")


class CompassView(QtWidgets.QMainWindow, _UI_compassview.Ui_MainWindow):
    '''
    Compass view window
    '''
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.b_bwd.setText('\u25C0')
        self.b_fwd.setText('\u25B6')
        self.b_cnt.setText('\u25CE')

        self.compassview = CompassWidget()
        self.gridLayout_compass.addWidget(self.compassview)
        pg.GraphicsView.setBackground(self.compassview, pg.mkColor(255, 255, 255, 255))
        # self.compassview.setStyleSheet("QWidget {padding: 20px; background-color: white;}")

        self.l_scale.setValidator(QDoubleValidator())
        self.l_mindepth.setValidator(QIntValidator())
        self.l_maxdepth.setValidator(QIntValidator())
        self.l_imheight.setValidator(QIntValidator())
        self.l_imwidth.setValidator(QIntValidator())

        self.b_bwd.clicked.connect(mc.navigate)
        self.b_fwd.clicked.connect(mc.navigate)
        self.b_cnt.clicked.connect(mc.navigate)
        self.b_SaveIm.clicked.connect(lambda: mc.saveimage('', 'one', 'compass'))
        self.b_SaveIms.clicked.connect(lambda: mc.saveallimages('compass'))
        self.l_scale.textEdited.connect(lambda: self.drawcompass(mc.col_to_show))
        self.l_mindepth.textEdited.connect(lambda: self.drawcompass(mc.col_to_show))
        self.l_maxdepth.textEdited.connect(lambda: self.drawcompass(mc.col_to_show))
        self.c_showlegend.stateChanged.connect(lambda: self.drawcompass(mc.col_to_show))
        self.c_showbt.stateChanged.connect(lambda: self.drawcompass(mc.col_to_show))
        self.l_legendtop_c.textEdited.connect(mc.changelegendtop)
        self.b_resize.clicked.connect(self.changesize)

        '''
        'view_win' parameter sets compassview view window no of neighbour profiles view_win = 2 means +-1 profile
        will be shown either side of the current one
        '''
        self.view_win = 2
        self.l_imheight.setText(str(self.compassview.height()))
        self.l_imwidth.setText(str(self.compassview.width()))


    def closeEvent(self, e):
        mc.c_showcompview.setChecked(False)


    def resizeEvent(self, e):
        self.l_imheight.setText(str(self.compassview.height()))
        self.l_imwidth.setText(str(self.compassview.width()))


    def changesize(self):
        self.showNormal()
        self.move(0, 0)
        self.resize(int(self.l_imwidth.text()) + 279, int(self.l_imheight.text()) + 28)


    def drawcompass(self, col_to_show):
        '''
        function shows compass view along with metadata (timestamp, Lat, Lon) on cv.compassview widget
        argument 'col_to_show' sets central profile as a column no of mc.t_data QTableWidget
        it uses 'cv.view_win' to set no o f profiles visible on the view
        '''
        self.compassview.clear()

        # set bckg white to avoid arrows leftovers
        pg.GraphicsView.setBackground(self.compassview, pg.mkColor(255, 255, 255, 255))

        self.setWindowTitle(f'Ensemble: {mc.META[mc.col_to_show, 0]}'
                            f' Time: {mc.META[mc.col_to_show, 1]}'
                            f' Lat: {mc.META[mc.col_to_show, 2]}'
                            f' Lon: {mc.META[mc.col_to_show, 3]}')

        meta_font = QFont()
        meta_font.setPixelSize(9)

        mc.ensemble_no = mc.META[col_to_show, 0]    # ensemble number to show

        scale = float(self.l_scale.text())          # scale of vectors to plot
        minz = int(self.l_mindepth.text())          # top to show on plot
        maxz = int(self.l_maxdepth.text())          # bottom to show on plot

        # ENSEMBLES (numbers!!!) range to show (+- from central ensemble)
        left = int(mc.META[col_to_show, 0] - self.view_win / 2 - 1)
        right = int(mc.META[col_to_show, 0] + self.view_win / 2 + 1)

        # mc.ADCP COLUMNS (numbers!!!) range to show (+- from central column)
        first = int(col_to_show - self.view_win / 2) if (int(col_to_show - self.view_win / 2) >= 0) else 0
        last = int(col_to_show + self.view_win / 2 + 1) if (int(col_to_show + self.view_win / 2 + 1) <= mc.no_ensembles)\
            else mc.no_ensembles

        # vertical (ensembles) and horisontal (level) infinite lines
        ens_lines, lev_lines = [], []
        for ens in range(left, right + 1):
            vline = pg.InfiniteLine(ens, angle=90, movable=False, pen=pg.mkPen('black', width=0.3,
                                                                               style=QtCore.Qt.DotLine))
            ens_lines.append(vline)
        for lev in -mc.ADCP[col_to_show, :, 2]:
            hline = pg.InfiniteLine(lev, angle=0, movable=False, pen=pg.mkPen('black', width=0.3,
                                                                              style=QtCore.Qt.DotLine))
            lev_lines.append(hline)
        # adding lines to plot
        for line in ens_lines + lev_lines:
            self.compassview.addItem(line)

        # !!!!!!!!!! FAST & FIXED WAY TO PLOT - first setParentItem, then addItem(parent)
        # vector arrows and metadata text
        parent_vec = pg.PlotDataItem()
        maxvec = np.max(mc.ADCP[first:last, :, 0])
        maxvec = 100 * math.ceil(maxvec / 100)
        for col in range(first, last):
            if self.c_showlegend.isChecked():
                # meta text
                meta = f'{str(mc.META[col, 1])}\nLat: {str(mc.META[col, 2])}\nLon: {str(mc.META[col, 3])}'
                meta_text = pg.TextItem(text=f'{meta}', color=(100, 100, 100, 255))
                meta_text.setFont(meta_font)
                meta_text.setPos(mc.META[col, 0], -minz)
                meta_text.setParentItem(parent_vec)
            # vector arrows
            for bin in range(mc.no_bins):
                arrow = InvertArrow(angle=mc.ADCP[col, bin, 1] + 90, headLen=scale * mc.ADCP[col, bin, 0], headWidth=1,
                                    pen=pg.mkPen(0, 128, 255, 255), brush='b')
                arrow.setPos(mc.META[col, 0], -mc.ADCP[0, bin, 2])
                arrow.setParentItem(parent_vec)

        # bottom track line
        if self.c_showbt.isChecked():
            x, y = mc.BTRA[:, 0].tolist(), (-mc.BTRA[:, 2]).tolist()
            btrack = pg.PlotDataItem(x=x, y=y, pen=pg.mkPen((0, 0, 0, 255), width=2))
            btrack.setParentItem(parent_vec)

        # add parent view to plot
        self.compassview.addItem(parent_vec)


        # scale arrow and text IMPORTANT TO self.compassview.plotItem.addItem(scale_arrow) & scale_arrow.setParentItem(parent_scale) together for updating items!!!!!
        parent_scale = self.compassview.plotItem
        scale_arrow = InvertArrow(angle=180, headLen=scale * maxvec, headWidth=1,
                                  pen=pg.mkPen(255, 0, 0, 255), brush='r')
        scale_arrow.setPos(100, 50)
        self.compassview.plotItem.addItem(scale_arrow)      # !!!!!!!
        scale_arrow.setParentItem(parent_scale)             # !!!!!!!

        scale_text = pg.TextItem(text=f'{maxvec} mm/s', color=(255, 0, 0, 255))
        scale_text.setPos(100, 50)
        self.compassview.plotItem.addItem(scale_text)       # !!!!!!!
        scale_text.setParentItem(parent_scale)              # !!!!!!!

        # set axes on compass view
        self.compassview.setRange(yRange=[-maxz - mc.bin_size, -minz + mc.bin_size])
        self.compassview.setRange(xRange=[left, right])
        self.compassview.getAxis('bottom').setLabel('Ensemble')
        self.compassview.getAxis('bottom').setTickSpacing(1, 1)
        self.compassview.getAxis('left').setTickSpacing(mc.bin_size, mc.bin_size)


class ProfileView(QtWidgets.QMainWindow, _UI_profview.Ui_MainWindow):
    '''
    Profile view window
    '''
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.b_bwd.setText('\u25C0')
        self.b_fwd.setText('\u25B6')
        if os.path.isfile(os.path.join(configfold,'top_view.png')):
            self.b_top.setIcon(QtGui.QIcon(os.path.join(configfold,'top_view.png')))
            self.b_top.setIconSize(QtCore.QSize(45, 45))
        if os.path.isfile(os.path.join(configfold,'se_view.png')):
            self.b_se.setIcon(QtGui.QIcon(os.path.join(configfold,'se_view.png')))
            self.b_se.setIconSize(QtCore.QSize(45, 45))


        self.b_bwd.clicked.connect(mc.navigate)
        self.b_fwd.clicked.connect(mc.navigate)
        self.b_top.clicked.connect(mc.navigate)
        self.b_se.clicked.connect(mc.navigate)
        self.b_SaveIm.clicked.connect(lambda: mc.saveimage('', 'one', 'profile'))
        self.b_SaveIms.clicked.connect(lambda: mc.saveallimages('profile'))
        self.c_showlabels.stateChanged.connect(lambda: self.drawADCP(mc.col_to_show))
        self.c_showrangecircle.stateChanged.connect(lambda: self.drawADCP(mc.col_to_show))
        self.c_showrangelabel.stateChanged.connect(lambda: self.drawADCP(mc.col_to_show))
        self.c_showgrid.stateChanged.connect(lambda: self.drawADCP(mc.col_to_show))
        self.c_showbottom.stateChanged.connect(lambda: self.drawADCP(mc.col_to_show))
        self.l_legendtop_p.textEdited.connect(mc.changelegendtop)

        self.profileview.setBackgroundColor(255, 255, 255, 255)
        self.profileview.setCameraPosition(pos=Vector(0.15, 0.1, -1), distance=2.5, elevation=45, azimuth=45)


    def closeEvent(self, e):
        mc.c_showprofview.setChecked(False)


    def drawADCP(self, col_to_show):
        '''
        function shows 3D profile view along with metadata (WD, magnitude, direction) on pv.profileview widget
        argument 'col_to_show' sets central profile as a column no of mc.t_data QTableWidget
        the plot is scaled to 1
        '''

        mc.col_to_show = col_to_show                        # !!! RE-SET value for multiple images saving (to use in saveimage func)
        mc.ensemble_no = mc.META[col_to_show, 0]            # ensemble number
        maxvec = np.max(mc.ADCP[mc.col_to_show, :, 0])      # max current vector in profile
        maxvec = 100 * math.ceil(maxvec / 100)              # max current vector in profile rounded to ceiling 100
        verscale = 1 / mc.no_bins                           # vertical scale
        horscale = 1 / maxvec                               # horizontal scale
        bottom = -(1 + verscale)                            # drawn level of seabed

        # clear plot
        self.profileview.clear()

        # items with a greater depth value will be drawn later
        # GLGraphicsItem cannot be set as parent to GLTextItem - thus GLTextItem added with addItem() method
        parent_prof = pg.opengl.GLGraphicsItem.GLGraphicsItem()

        # grid
        grid = gl.GLGridItem(glOptions='additive')
        grid.setColor(QColor(200, 200, 200, 255))
        grid.setSize(2, 2, 0)
        grid.setSpacing(0.25, 0.25, 0)
        grid.translate(0, 0, -(1 + verscale))
        grid.setDepthValue(1)

        # floor
        # verts = np.array([[-1, -1, 1.001 * bottom], [1, -1, 1.001 * bottom],
        #                   [1, 1, 1.001 * bottom], [-1, 1, 1.001 * bottom]])
        # faces = np.array([[0, 1, 2], [0, 2, 3]])

        verts = np.array([[-1, -1, bottom], [1, -1, bottom],
                          [1, 1, bottom], [-1, 1, 1.001 * bottom],
                          [-0.75, -0.03, bottom], [-0.9, 0, bottom],
                          [-0.75, 0.03, bottom]])                    # encompass north arrow to avoid rendering mixture
        faces = np.array([[0, 1, 4], [0, 4, 5], [0, 5, 3], [3, 5, 6],
                          [6, 2, 3], [6, 2, 4], [2, 4, 1]])
        primitive_seabed = gl.MeshData(vertexes=verts, faces=faces)
        seabed = gl.GLMeshItem(meshdata=primitive_seabed, smooth=True, color=QColor(240, 240, 240, 255),
                               glOptions='translucent')

        seabed.setDepthValue(-10)

        # plumb-line
        primitive_plumb = gl.MeshData.cylinder(rows=10, cols=25, radius=[0.0025, 0.0025], length=(1 + 2 * verscale),
                                               offset=False)
        plumb = gl.GLMeshItem(meshdata=primitive_plumb, smooth=True, color=QColor(0, 0, 0, 255),
                              glOptions='translucent')
        plumb.translate(0, 0, bottom)
        plumb.setDepthValue(1)

        # north line
        primitive_nline = np.array([[0, 0, bottom], [-0.75, 0, bottom]])
        n_line = gl.GLLinePlotItem(pos=primitive_nline, width=2, color=QColor(255, 0, 0, 255), antialias=True,
                                   glOptions='additive')
        n_line.setDepthValue(1)

        # north arrow
        vert_narrow = np.array([[-0.75, -0.03, bottom],
                                [-0.9, 0, bottom],
                                [-0.75, 0.03, bottom]])
        face_narrow = np.array([[0, 1, 2]])
        n_arrow = gl.GLMeshItem(vertexes=vert_narrow, faces=face_narrow, color=QColor(255, 0, 0, 255),
                                smooth=False, glOptions='translucent')


        # primitive_narrow = np.array([[-0.75, -0.03, bottom], [-0.9, 0, bottom],
        #                              [-0.75, 0.03, bottom], [-0.75, -0.03, bottom]])
        # n_arrow = gl.GLLinePlotItem(pos=primitive_narrow, width=2, color=QColor(255, 0, 0, 255), antialias=True,
        #                             glOptions='additive')
        n_arrow.setDepthValue(10)

        # north mark
        primitive_n = np.array([[-0.9, -0.0175, bottom], [-1, -0.0175, bottom],
                                [-0.9, 0.0175, bottom], [-1, 0.0175, bottom]])
        n_mark = gl.GLLinePlotItem(pos=primitive_n, width=2, color=QColor(255, 0, 0, 255), antialias=True,
                                   glOptions='additive')
        n_mark.setDepthValue(1)

        # range circle
        primitive_be = np.stack((np.cos(np.linspace(0, 2 * np.pi, 360)),
                                 np.sin(np.linspace(0, 2 * np.pi, 360)),
                                 np.full(360, bottom)), axis=1)
        be = gl.GLLinePlotItem(pos=primitive_be, width=1, color=QColor(0, 0, 255, 255), antialias=True,
                               glOptions='additive')
        be.setDepthValue(1)

        # range label
        be_label = gl.GLTextItem(pos=[0, 1, bottom], text=f'{str(maxvec)} mm/s', color=(0, 0, 255, 255),
                                 font=QFont('Arial', 14))
        be_label.setDepthValue(1)

        # vectors and labels
        vec_labs = []
        for b in range(mc.no_bins):
            # current vectors as cylinders
            primitive_vec = gl.MeshData.cylinder(rows=10, cols=25, radius=[0.005, 0.005],
                                                 length=mc.ADCP[mc.col_to_show, b, 0] * horscale, offset=False)
            vec = gl.GLMeshItem(meshdata=primitive_vec, smooth=True, color=QColor(0, 128, 255, 255),
                                glOptions='translucent')
            vec.rotate(90, 0, -1, 0)
            vec.translate(0, 0, -mc.ADCP[col_to_show, b, 3] * verscale)
            vec.rotate(-mc.ADCP[mc.col_to_show, b, 1], 0, 0, 1)
            vec.setDepthValue(2)
            vec.setParentItem(parent_prof)

            # current dept/magnitude/direction as text
            lab = gl.GLTextItem(pos=[-mc.ADCP[mc.col_to_show, b, 0] * horscale, 0, -mc.ADCP[mc.col_to_show, b, 3] * verscale],
                                text=f'{str(mc.ADCP[mc.col_to_show, b, 2])}m / {str(int(mc.ADCP[mc.col_to_show, b, 0]))}mm/s / '
                                     f'{str(mc.ADCP[mc.col_to_show, b, 1])}\u00B0',
                                color=(0, 0, 255, 255), font=QFont('Arial', 8))
            lab.rotate(mc.ADCP[mc.col_to_show, b, 1], 0, 0, -1)
            vec_labs.append(lab)

        # adding elements to plot
        if self.c_showbottom.isChecked():
            seabed.setParentItem(parent_prof)
        if self.c_showgrid.isChecked():
            grid.setParentItem(parent_prof)
        if self.c_showrangecircle.isChecked():
            be.setParentItem(parent_prof)
        if self.c_showrangelabel.isChecked():
            self.profileview.addItem(be_label)

        plumb.setParentItem(parent_prof)
        n_line.setParentItem(parent_prof)
        n_mark.setParentItem(parent_prof)
        n_arrow.setParentItem(parent_prof)

        self.profileview.addItem(parent_prof)

        if self.c_showlabels.isChecked():
            for item in vec_labs:
                self.profileview.addItem(item)


class GeoView(QtWidgets.QMainWindow, _UI_geoview.Ui_MainWindow):
    '''
    Geographical view window
    '''
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.geoview.setAspectLocked(lock=True, ratio=2)
        pg.GraphicsView.setBackground(self.geoview, pg.mkColor(255, 255, 255, 255))
        self.geoview.showGrid(x=True, y=True)
        self.geoview.showAxis('top')
        self.geoview.showAxis('right')
        self.geoview.plotItem.setMenuEnabled(False)                     # Disable context menu

        self.l_scale.setValidator(QDoubleValidator())

        self.sl_bin.valueChanged.connect(self.drawgeo)
        self.l_scale.textChanged.connect(self.drawgeo)
        self.b_SaveIm.clicked.connect(lambda: mc.saveimage('', 'one', 'map'))


    def closeEvent(self, e):
        mc.c_showmapview.setChecked(False)


    def drawgeo(self):
        self.geoview.clear()


        meta_font = QFont()
        meta_font.setPixelSize(10)

        bin = -self.sl_bin.value()
        scale = float(self.l_scale.text())  # scale of vectors to plot

        if self.c_colorcodedepth.isChecked():
            pg.GraphicsView.setBackground(self.geoview, pg.mkColor(int(255 * (1 - bin / (mc.maxbin + 1))),
                                                                   255, 255, 255))
        else:
            # sets bckg white to avoid arrows leftovers from changing scale
            pg.GraphicsView.setBackground(self.geoview, pg.mkColor(255, 255, 255, 255))

        parent_geo = pg.PlotDataItem()
        for i in range(len(mc.fNames)):
            # track
            x, y = mc.META_s[i][:, 3].tolist(), mc.META_s[i][:, 2].tolist()
            track = pg.PlotDataItem(x=x, y=y, pen=pg.mkPen((255, 0, 0, 255), width=1), symbol='o', symbolPen=None, symbolSize=4)
            track.setParentItem(parent_geo)

            # linename label
            meta_label = f'{mc.fNames_list[i]}\n'

            # try to add WD - fails if max bin of profile is less than maxbin
            try:
                meta_label += f'WD: {mc.ADCP_s[i][0, bin, 2]}m'
            except:
                pass

            meta_text = pg.TextItem(text=meta_label, color=(0, 0, 0, 255))
            meta_text.setFont(meta_font)
            meta_text.setPos(mc.META_s[i][0, 3], mc.META_s[i][0, 2])
            meta_text.setParentItem(parent_geo)

            for j in range(len(mc.META_s[i])):
                try:
                    arrow = InvertArrow(angle=mc.ADCP_s[i][j, bin, 1] + 90, headLen=scale * mc.ADCP_s[i][j, bin, 0], headWidth=1,
                                        pen=pg.mkPen(0, 128, 255, 255), brush='b')
                    arrow.setPos(mc.META_s[i][j, 3], mc.META_s[i][j, 2])
                    arrow.setParentItem(parent_geo)
                except:
                    pass

        self.geoview.addItem(parent_geo)

        parent_scale = self.geoview.plotItem
        scale_arrow = InvertArrow(angle=180, headLen=scale * mc.maxvec, headWidth=1,
                                  pen=pg.mkPen(255, 0, 0, 255), brush='r')
        scale_arrow.setPos(100, 50)
        self.geoview.plotItem.addItem(scale_arrow)      # !!!!!!!
        scale_arrow.setParentItem(parent_scale)             # !!!!!!!

        scale_text = pg.TextItem(text=f'{mc.maxvec} mm/s', color=(255, 0, 0, 255))
        scale_text.setPos(100, 50)
        self.geoview.plotItem.addItem(scale_text)       # !!!!!!!
        scale_text.setParentItem(parent_scale)              # !!!!!!!


class MainWindow(QtWidgets.QMainWindow, _UI_control_3D_Geo.Ui_MainWindow):
    '''
    Main window
    '''
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.t_data.horizontalHeader().setStyleSheet("::section{border: 1px solid gray;}")
        self.t_data.verticalHeader().setStyleSheet("::section{border: 1px solid gray;}")
        self.t_data.setSelectionBehavior(QAbstractItemView.SelectColumns)
        self.t_data.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.b_open.clicked.connect(self.openfiles)
        self.c_filenames.currentIndexChanged.connect(self.selectADCPfile)
        self.b_exportxl.clicked.connect(lambda: self.exportxl('one'))
        self.b_exportxlall.clicked.connect(lambda: self.exportxl('all'))
        self.b_exportgis.clicked.connect(self.exportgis)
        self.t_data.cellDoubleClicked.connect(self.showensemble)
        self.c_showprofview.stateChanged.connect(self.showviews)
        self.c_showcompview.stateChanged.connect(self.showviews)
        self.c_showmapview.stateChanged.connect(self.showviews)
        self.actionLicense.triggered.connect(lambda: li.show())
        self.actionHow_to_use.triggered.connect(lambda: he.show())

        # used for iterating through ensembles when drawing plots (without showing) for images
        self.emsemble_no = 0

        # self.showviews()


    def closeEvent(self, e):
        pv.close()
        cv.close()
        gv.close()
        li.close()
        he.close()


    def showviews(self):
        if self.c_showprofview.isChecked():
            pv.show()
        else:
            pv.close()

        if self.c_showcompview.isChecked():
            cv.show()
        else:
            cv.close()

        if self.c_showmapview.isChecked():
            gv.show()
        else:
            gv.close()


    def openfiles(self):
        '''
        creates list of mat files (this does not read mat files)
        - passes control to self.selectADCPfile
        '''
        self.fNames, _ = QFileDialog.getOpenFileNames(self, 'Load ADCP data', '',
                                               'data file (*.mat);;All Files (*)', options=OPTIONS)

        if self.fNames:
            self.c_filenames.clear()
            self.loadADCPdata()


    def loadADCPdata(self):
        '''
        func to load all ADCP files into lists
        '''
        if self.fNames:
            self.fNames_list = []               # list of loaded files
            fNames_rejected_list = []           # list of corrupted files to reject from self.fNames

            self.maxbin, self.minLat, self.maxLat, self.minLon, self.maxLon, self.maxvec = (
                0, 91, 0, 181, 0, 0)
            self.adcp_mat_s = []
            (self.fileheader_s, self.no_ensembles_s, self.no_bins_s,
             self.bin_size_s, self.ADCP_s, self.META_s, self.BTRA_s) = (
                [], [], [], [], [], [], [])
            for fName in self.fNames:
                try:
                    adcp_mat = _read_matlab1_AK.load_mat(fName)
                    (fileheader, no_ensembles, no_bins, bin_size, ADCP, META, BTRA) = _fill_ADCP_mx.fill_mx(adcp_mat)

                    maxbin = np.max(no_bins)
                    minLat = np.min(META[:, 2])
                    maxLat = np.max(META[:, 2])
                    minLon = np.min(META[:, 3])
                    maxLon = np.max(META[:, 3])
                    maxvec = np.max(ADCP[:, :, 0])

                    self.maxbin = maxbin if maxbin > self.maxbin else self.maxbin
                    self.minLat = minLat if minLat < self.minLat else self.minLat
                    self.maxLat = maxLat if maxLat > self.maxLat else self.maxLat
                    self.minLon = minLon if minLon < self.minLon else self.minLon
                    self.maxLon = maxLon if maxLon > self.maxLon else self.maxLon
                    self.maxvec = maxvec if maxvec > self.maxvec else self.maxvec
                    self.maxvec = 100 * math.ceil(self.maxvec / 100)

                    self.adcp_mat_s.append(adcp_mat)

                    self.fileheader_s.append(fileheader)
                    self.no_ensembles_s.append(no_ensembles)
                    self.no_bins_s.append(no_bins)
                    self.bin_size_s.append(bin_size)
                    self.ADCP_s.append(ADCP)
                    self.META_s.append(META)
                    self.BTRA_s.append(BTRA)

                    self.fNames_list.append(Path(fName).stem)

                except:
                    fNames_rejected_list.append(fName)
                    messagepop(f'File {fName} corrupted or data fields missing\nFile will not be loaded')

            # reject corrupted filnames from self.fNames
            for fName_rejected in fNames_rejected_list:
                self.fNames.remove(fName_rejected)

            self.meanLat = (self.minLat + self.maxLat) / 2

            # set geo X/Y -> Lat/Lon scale at projection
            gv.geoview.setAspectLocked(lock=True, ratio=math.cos(math.radians(self.meanLat)))
            # set geo slider max to max bin
            gv.sl_bin.setRange(-self.maxbin - 1, 0)

            self.c_filenames.addItems(self.fNames_list)

            gv.drawgeo()
            self.selectADCPfile()


    def selectADCPfile(self):
        '''
        func loads selected ADCP file using '_read_matlab1_AK' and '_fill_ADCP_mx'
        - passses control to self.showensemble
        '''
        currentfileix = self.c_filenames.currentIndex()
        fName = self.fNames[currentfileix]

        if fName:
            try:
                self.fileheader = self.fileheader_s[currentfileix]
                self.no_ensembles = self.no_ensembles_s[currentfileix]
                self.no_bins = self.no_bins_s[currentfileix]
                self.bin_size = self.bin_size_s[currentfileix]
                self.ADCP = self.ADCP_s[currentfileix]
                self.META = self.META_s[currentfileix]
                self.BTRA = self.BTRA_s[currentfileix]

                self.fill_ADCP_table()

                self.setWindowTitle(f'ADCPlot 3D Geo: {os.path.basename(fName)}')
                self.ADCPfname = fName

                pv.l_legendtop_p.setText(Path(self.ADCPfname).stem)
                cv.l_legendtop_c.setText(Path(self.ADCPfname).stem)

                cv.l_mindepth.setText('0')
                cv.l_maxdepth.setText(str(int(np.max(mc.ADCP[0, :, 2]))))

                self.showensemble(0, 0)

                # set geo plot extents
                minLat, maxLat = np.min(self.META[:, 2]), np.max(self.META[:, 2])
                minLon, maxLon = np.min(self.META[:, 3]), np.max(self.META[:, 3])
                gv.geoview.setRange(yRange=[minLat, maxLat])
                gv.geoview.setRange(xRange=[minLon, maxLon])

            except:
                messagepop('File corrupted or data fields missing')


    def fill_ADCP_table(self):
        '''
        func to load ADCP data into mc header and table
        '''
        self.txt_header.clear()
        for k, v in self.fileheader.items():
            self.txt_header.append(f'{k}: {str(v).replace('[', '').replace(']', '')}')

        self.t_data.setColumnCount(self.no_ensembles)
        self.t_data.setRowCount(self.no_bins)
        self.t_data.setHorizontalHeaderLabels(self.META[:, 0].astype('int').astype('str'))  # ensemble no
        self.t_data.setVerticalHeaderLabels(self.ADCP[0, :, 2].astype('str') + 'm')         # bin level

        for table_row in range(self.no_bins):
            for table_col in range(self.no_ensembles):
                cell_ent = f'M: {str(self.ADCP[table_col, table_row, 0])}\nD: {str(self.ADCP[table_col, table_row, 1])}'

                self.t_data.setItem(table_row, table_col, QtWidgets.QTableWidgetItem(cell_ent))


    def showensemble(self, row, column):
        '''
        func to set views for selected profile
        this is called after double-clicking table on mc, loading ADCP file or navigating throug profiles by buttons
        '''
        self.col_to_show = column

        pv.setWindowTitle(f'Ensemble: {self.META[self.col_to_show, 0]}'
                          f' Time: {self.META[self.col_to_show, 1]}'
                          f' Lat: {self.META[self.col_to_show, 2]}'
                          f' Lon: {self.META[self.col_to_show, 3]}')
        pv.drawADCP(self.col_to_show)

        if cv.l_mindepth.text() == '' or cv.l_maxdepth.text() == '':        # first time shown
            cv.l_mindepth.setText('0')
            cv.l_maxdepth.setText(str(int(np.max(self.ADCP[self.col_to_show, :, 2]))))
        cv.drawcompass(self.col_to_show)


    def navigate(self):
        '''
        func navigates through profiles in the loaded file using buttons
        - passses control to self.showensemble
        '''
        sender = self.sender().objectName()

        if sender == 'b_bwd':
            col_to_show = mc.col_to_show - 1 if mc.col_to_show >= 1 else 0
        if sender == 'b_fwd':
            col_to_show = mc.col_to_show + 1 if mc.col_to_show < mc.no_ensembles - 1 else mc.no_ensembles - 1
        if sender == 'b_cnt':
            col_to_show = int(self.no_ensembles / 2)
            cv.view_win = int(self.no_ensembles / 2) * 2
            cv.l_maxdepth.setText(str(math.ceil(np.max(self.BTRA[:, 2]))))
        if sender == 'b_se':
            pv.profileview.setCameraPosition(pos=Vector(0.15, 0.1, -1), distance=2.5, elevation=45, azimuth=45)
            col_to_show = mc.col_to_show
            pv.profileview.opts['fov'] = 60
        if sender == 'b_top':
            pv.profileview.setCameraPosition(pos=Vector(0, 0, 0), distance=175, elevation=90, azimuth=0)
            pv.profileview.opts['fov'] = 1
            col_to_show = mc.col_to_show

        self.showensemble(0, col_to_show)


    def changelegendtop(self):
        '''
        func to change text in legend top text entries in pv and cv
        '''
        sender = self.sender().objectName()

        if sender == 'l_legendtop_p':
            cv.l_legendtop_c.setText(pv.l_legendtop_p.text())
        else:
            pv.l_legendtop_p.setText(cv.l_legendtop_c.text())


    def saveallimages(self, plottype):
        '''
        func to specify parameters of multiple images save
        - passes control to self.saveimage
        '''
        foldName = QFileDialog.getExistingDirectory(self, 'Select folder to save all images', '', options=OPTIONS)
        if foldName:
            if plottype == 'profile':
                step = 1
                start = 0
                end = mc.no_ensembles
            if plottype == 'compass':
                step = 1 if cv.view_win == 0 else cv.view_win
                start = int(step / 2)
                end = mc.no_ensembles

            count = 0
            for i in range(start, end, step):
                if plottype == 'profile':
                    pv.drawADCP(i)
                if plottype == 'compass':
                    cv.drawcompass(i)
                self.saveimage(foldName, 'all', plottype)
                count += 1

            messagepop(f'{count} Images saved')


    def saveimage(self, savefolder, what_to_save, plottype):
        '''
        func to save image
        '''
        if what_to_save == 'all':
            imagename = os.path.join(savefolder, Path(self.ADCPfname).stem + '_' + plottype + '_' + str(self.ensemble_no) + '.png')

        if what_to_save == 'one':
            fName, _ = QFileDialog.getSaveFileName(self, 'Save this image', '*',
                                                   'png file (*.png);;All Files (*)', options=OPTIONS)
            imagename = fName


        if plottype == 'compass':
            # export PlotItem with exporter
            exporter = pg.exporters.ImageExporter(cv.compassview.scene())
            exporter.export(imagename)
            # open image with PIL and add legend
            image = Image.open(imagename)
            draw = ImageDraw.Draw(image)
            font = ImageFont.truetype('arial', 16)
            draw.text((40, 10), f'{pv.l_legendtop_p.text()}\n',
                      fill=(0, 0, 0), font=font)
            image.save(imagename)

        if plottype == 'profile':
            # save openGL 3D with grabFramebuffer
            pv.profileview.grabFramebuffer().save(imagename)
            # open image with PIL and add legend
            image = Image.open(imagename)
            draw = ImageDraw.Draw(image)
            font = ImageFont.truetype('arial', 16)
            draw.text((10, 10), f'{pv.l_legendtop_p.text()}\n'
                                f'Ensemble: {self.ensemble_no}\n'
                                f'Time: {self.META[self.col_to_show, 1]}\n'
                                f'Lat: {self.META[self.col_to_show, 2]}\n'
                                f'Lon: {self.META[self.col_to_show, 3]}',
                      fill=(0, 0, 0), font=font)
            image.save(imagename)

        if plottype == 'map':
            # export PlotItem with exporter
            exporter = pg.exporters.ImageExporter(gv.geoview.scene())
            exporter.export(imagename)
            # open image with PIL and add legend
            image = Image.open(imagename)
            draw = ImageDraw.Draw(image)
            font = ImageFont.truetype('arial', 16)
            draw.text((40, 20), f'{gv.l_legendtop_g.text()}\n',
                      fill=(0, 0, 0), font=font)
            image.save(imagename)


        if what_to_save == 'one':
            messagepop('Image saved')


    def exportxl(self, what_to_save):
        '''
        func to export loaded files to Excel (all available)
        '''
        if what_to_save == 'one':
            savefName, _ = QFileDialog.getSaveFileName(self, 'Export to Excel', '',
                                                       'Excel file (*.xlsx);;All Files (*)', options=OPTIONS)
            fNames = [self.ADCPfname]

        if what_to_save == 'all':
            savefoldName = QFileDialog.getExistingDirectory(self, 'Select folder to save all files', '',
                                                            options=OPTIONS)
            fNames = self.fNames

        if fNames:
            count = 0
            for f, fName in enumerate(fNames):
                if what_to_save == 'one':
                    f = self.c_filenames.currentIndex()
                if what_to_save == 'all':
                    savefName = os.path.join(savefoldName, self.fNames_list[f] + '.xlsx')

                xlwb = Workbook()
                xlws = xlwb.create_sheet('FILE HEADER')

                # 'FILE HEADER'
                i = 1
                for k, v in self.fileheader_s[f].items():
                    xlws.cell(row=i, column=1).value = k
                    xlws.cell(row=i, column=2).value = str(v).replace('[', '').replace(']', '')
                    i += 1

                sheets_S, sheets_A = [], []
                for k, value in self.adcp_mat_s[f].items():
                    if 'Ser' in k and k not in ['SerBins', 'SerEnsembles', 'SerYear', 'SerMon', 'SerDay',
                                                'SerHour', 'SerMin', 'SerSec', 'SerHund']:
                        sheets_S.append(k)
                    if k[:2] == 'An':
                        sheets_A.append(k)

                for sheet in sheets_S:
                    xlws = xlwb.create_sheet(sheet)
                    for i, ensemble in enumerate(self.META_s[f]):
                        xlws.cell(row=1, column=i + 3).value = ensemble[0]
                        xlws.cell(row=2, column=i + 3).value = ensemble[1]
                        xlws.cell(row=3, column=i + 3).value = ensemble[2]
                        xlws.cell(row=4, column=i + 3).value = ensemble[3]
                    for j, bin in enumerate(self.ADCP_s[f][0]):
                        xlws.cell(row=j + 5, column=1).value = bin[3]
                        xlws.cell(row=j + 5, column=2).value = bin[2]

                    for i in range(self.no_ensembles_s[f]):
                        for j in range(self.no_bins_s[f]):
                            xlws.cell(row=j + 5, column=i + 3).value = self.adcp_mat_s[f][sheet].T[i, j]

                for sheet in sheets_A:
                    xlws = xlwb.create_sheet(sheet)
                    for i, ensemble in enumerate(self.META_s[f]):
                        xlws.cell(row=1, column=i + 1).value = ensemble[0]
                        xlws.cell(row=2, column=i + 1).value = ensemble[1]
                        xlws.cell(row=3, column=i + 1).value = ensemble[2]
                        xlws.cell(row=4, column=i + 1).value = ensemble[3]

                    for i in range(self.no_ensembles_s[f]):
                            xlws.cell(row=5, column=i + 1).value = self.adcp_mat_s[f][sheet].T[i][0]


                xlwb.save(savefName)
                count += 1

            messagepop(f'{count} Excel file(s) saved')


    def exportgis(self):
        '''
        func to export all selected files to GIS compatible GEOJSON/WKT
        '''
        outfname, exttype = QFileDialog.getSaveFileName(self,
                                                        'Save As', '*',
                                                        'geojson files (*.geojson);;wkt files (*.wkt);;all files (*.*)',
                                                        options=OPTIONS)

        if outfname and self.fNames:
            if 'wkt' in exttype:
                gisfiletype = '.wkt'            
            if 'geojson' in exttype:
                gisfiletype = '.geojson'
                
            # all bins
            outfname_all = os.path.join(os.path.dirname(outfname),
                                        f'{Path(outfname).stem}{gisfiletype}')
            # surface bin (bin == 0)
            outfname_sur = os.path.join(os.path.dirname(outfname),
                                        f'{Path(outfname).stem}_surface{gisfiletype}')
            # bottom bin (last non-zero bin)
            outfname_bot = os.path.join(os.path.dirname(outfname),
                                        f'{Path(outfname).stem}_bottom{gisfiletype}')

            if gisfiletype == '.wkt':
                # wkt header
                wkt_str_a = f'geometry;Lon;Lat;WD;Timestamp;T_Name;Ens_Name;Bin_No;Magnitude;Direction\n'
                wkt_str_s = f'geometry;Lon;Lat;WD;Timestamp;T_Name;Ens_Name;Bin_No;Magnitude;Direction\n'
                wkt_str_b = f'geometry;Lon;Lat;WD;Timestamp;T_Name;Ens_Name;Bin_No;Magnitude;Direction\n'

            if gisfiletype == '.geojson':
                # json dict
                JSONDICT_A = {
                    'type': 'FeatureCollection',
                    'name': Path(outfname).stem,
                    # 'crs': {
                    #     'type': 'name',
                    #     'properties': {
                    #         'name': 'urn:ogc:def:crs:OGC:1.3:CRS84'
                    #     }
                    # },
                }
                JSONDICT_S = {
                    'type': 'FeatureCollection',
                    'name': Path(outfname).stem,
                }
                JSONDICT_B = {
                    'type': 'FeatureCollection',
                    'name': Path(outfname).stem,
                }                

                featureslist_a, featureslist_s, featureslist_b  = [], [], []

            count = 1
            for i in range(len(self.fNames)):
                for ens in range(self.no_ensembles_s[i]):
                    for bin in range(self.no_bins_s[i]):                        
                        if gisfiletype == '.wkt':
                            # all bins
                            singlefeaturestring = (f'POINT({self.META_s[i][ens, 3]} {self.META_s[i][ens, 2]}'
                                             f' -{self.ADCP_s[i][ens, bin, 2]});'
                                             f'{self.META_s[i][ens, 3]};{self.META_s[i][ens, 2]};{self.ADCP_s[i][ens, bin, 2]};'
                                             f'{self.META_s[i][ens, 1]};{Path(self.fNames[i]).stem};{self.META_s[i][ens, 0]};'
                                             f'{self.ADCP_s[i][ens, bin, 3]};'
                                             f'{self.ADCP_s[i][ens, bin, 0]};{self.ADCP_s[i][ens, bin, 1]}\n')                           
                            wkt_str_a += singlefeaturestring                # adds every feature                          
                            # surface bin
                            if bin == 0:
                                wkt_str_s += singlefeaturestring            # adds first bin feature
                            # bottom bin
                            if self.ADCP_s[i][ens, bin, 0] != 0 and self.ADCP_s[i][ens, bin, 1] != 0:                                                                                              
                                singlefeaturestring_b = singlefeaturestring # overrdes last non-zero bin feature


                        if gisfiletype == '.geojson':
                            singlefeaturedict = {
                                'type': 'Feature',
                                'properties': {
                                    'fid': count,
                                    'Lon': self.META_s[i][ens, 3],
                                    'Lat': self.META_s[i][ens, 2],
                                    'WD': self.ADCP_s[i][ens, bin, 2],
                                    'Timestamp': self.META_s[i][ens, 1],
                                    'T_Name': Path(self.fNames[i]).stem,
                                    'Ens_Name': self.META_s[i][ens, 0],
                                    'Bin_No': self.ADCP_s[i][ens, bin, 3],
                                    'Magnitude': self.ADCP_s[i][ens, bin, 0],
                                    'Direction': self.ADCP_s[i][ens, bin, 1],
                                },
                                'geometry': {
                                    'type': 'Point',
                                    'coordinates': [self.META_s[i][ens, 3], self.META_s[i][ens, 2], -self.ADCP_s[i][ens, bin, 2]],
                                }
                            }

                            count += 1

                            # all ins
                            featureslist_a.append(singlefeaturedict)        # adds every feature                           
                            # surface bin
                            if bin == 0:
                                featureslist_s.append(singlefeaturedict)    # adds first bin feature
                            # bottom bin
                            if self.ADCP_s[i][ens, bin, 0] != 0 and self.ADCP_s[i][ens, bin, 1] != 0:
                                singlefeaturedict_b = singlefeaturedict     # overrdes last non-zero bin feature
                  
                    if gisfiletype == '.wkt':
                        wkt_str_b += singlefeaturestring_b                    
                    if gisfiletype == '.geojson':
                        featureslist_b.append(singlefeaturedict_b)

                            
            if gisfiletype == '.wkt':
                for f, s in zip([outfname_all, outfname_sur, outfname_bot],
                                [wkt_str_a, wkt_str_s, wkt_str_b]):
                    with open(f, 'w') as outfile:
                        outfile.write(s)

            if gisfiletype == '.geojson':
                for f, d, l in zip([outfname_all, outfname_sur, outfname_bot],
                                   [JSONDICT_A, JSONDICT_S, JSONDICT_B],
                                   [featureslist_a, featureslist_s, featureslist_b]):
                    d['features'] = l
                    json_str = json.dumps(d, indent=0)
                    with open(f, 'w') as outfile:
                        outfile.write(json_str)


            messagepop(f'{gisfiletype[1:]} files created')


def messagepop(message):
    '''
    messagebox
    '''
    msg = QMessageBox()
    msg.setWindowTitle('Warning')
    msg.setText(message)
    if iconhere:
        msg.setWindowIcon(icon)
    msg.show()
    msg.exec()


def main():
    global mc
    global pv
    global cv
    global gv
    global li
    global he
    global icon
    global iconhere
    global configfold

    iconhere = False

    parentfold = os.getcwd()
    configfold = os.path.join(parentfold, '_internal')

    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('fusion')
    SCREEN_WIDTH, SCREEN_HEIGHT = app.primaryScreen().size().toTuple()

    mc = MainWindow()
    pv = ProfileView()
    cv = CompassView()
    gv = GeoView()
    li = LicenseWin()
    he = HelpWin()

    if os.path.isfile(os.path.join(configfold, 'blob.ico')):
        iconhere = True
        icon = QtGui.QIcon(os.path.join(configfold, 'blob.ico'))
        for win in [mc, pv, cv, gv, li, he]:
            win.setWindowIcon(icon)

    mc.setWindowTitle('ADCPlot 3D Geo')
    pv.setWindowTitle('Profile')
    cv.setWindowTitle('Compass')
    gv.setWindowTitle('Map')

    mc.setGeometry(0, 0, int(SCREEN_WIDTH / 2), int(SCREEN_HEIGHT / 2))
    pv.move(int(SCREEN_WIDTH / 2), 0)
    cv.move(int(SCREEN_WIDTH / 2), int(SCREEN_HEIGHT / 2))
    gv.move(int(SCREEN_WIDTH / 2), int(SCREEN_HEIGHT / 2))

    mc.show()
    cv.show()
    pv.show()
    # gv.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()